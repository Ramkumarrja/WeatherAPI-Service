import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from datetime import datetime

class ExcelService:
    def __init__(self):
        pass
    
    def generate_excel(self, weather_data):
        """Generate Excel file with weather data for last 48 hours"""
        if not weather_data:
            return self._generate_empty_excel()
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Weather Data"
        
        # Set up headers
        headers = ['timestamp', 'temperature_2m', 'relative_humidity_2m']
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for data in weather_data:
            row_data = [
                data.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                data.temperature if data.temperature is not None else 'N/A',
                data.humidity if data.humidity is not None else 'N/A'
            ]
            ws.append(row_data)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add metadata sheet
        metadata_ws = wb.create_sheet("Metadata")
        self._add_metadata_sheet(metadata_ws, weather_data)
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _add_metadata_sheet(self, ws, weather_data):
        """Add metadata information to a separate sheet"""
        if not weather_data:
            ws.append(["No data available"])
            return
        
        sample = weather_data[0]
        first_date = weather_data[0].timestamp.strftime('%Y-%m-%d %H:%M:%S')
        last_date = weather_data[-1].timestamp.strftime('%Y-%m-%d %H:%M:%S')
        
        # Count available data points
        temp_count = sum(1 for d in weather_data if d.temperature is not None)
        humidity_count = sum(1 for d in weather_data if d.humidity is not None)
        
        # Calculate ranges
        valid_temps = [d.temperature for d in weather_data if d.temperature is not None]
        valid_humidities = [d.humidity for d in weather_data if d.humidity is not None]
        
        metadata = [
            ["Report Information", ""],
            ["Generated At", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["", ""],
            ["Location Information", ""],
            ["Latitude", sample.latitude],
            ["Longitude", sample.longitude],
            ["", ""],
            ["Data Range", ""],
            ["Start Date", first_date],
            ["End Date", last_date],
            ["Total Records", len(weather_data)],
            ["", ""],
            ["Data Quality", ""],
            ["Temperature Records", temp_count],
            ["Humidity Records", humidity_count],
            ["", ""],
            ["Statistics", ""],
        ]
        
        if valid_temps:
            metadata.extend([
                ["Min Temperature (°C)", f"{min(valid_temps):.1f}"],
                ["Max Temperature (°C)", f"{max(valid_temps):.1f}"],
                ["Avg Temperature (°C)", f"{sum(valid_temps)/len(valid_temps):.1f}"],
            ])
        
        if valid_humidities:
            metadata.extend([
                ["Min Humidity (%)", f"{min(valid_humidities):.1f}"],
                ["Max Humidity (%)", f"{max(valid_humidities):.1f}"],
                ["Avg Humidity (%)", f"{sum(valid_humidities)/len(valid_humidities):.1f}"],
            ])
        
        metadata.extend([
            ["", ""],
            ["Data Source", "Open-Meteo MeteoSwiss API"],
            ["Data Type", "Historical (Past 2 Days)" if not getattr(sample, 'is_forecast', True) else "Forecast"]
        ])
        
        # Add metadata to sheet
        for row_data in metadata:
            ws.append(row_data)
        
        # Style the metadata sheet
        for row in ws.iter_rows():
            for cell in row:
                if cell.column == 1 and cell.value and cell.value.endswith("Information"):
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                elif cell.column == 1 and cell.value:
                    cell.font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _generate_empty_excel(self):
        """Generate Excel file when no data is available"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Weather Data"
        
        # Add headers
        headers = ['timestamp', 'temperature_2m', 'relative_humidity_2m']
        ws.append(headers)
        
        # Add empty row with message
        ws.append(['No data available for the selected time period', '', ''])
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num in range(1, 4):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
